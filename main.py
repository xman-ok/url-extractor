import os
import re
import time
from tkinter import filedialog, messagebox
import tkinter as tk
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# 웹 크롤링 표준 라이브러리
from selenium import webdriver
from selenium.webdriver.common.by import By


class LinkExtractorApp:

    def __init__(self, root):
        self.root = root
        self.root.title("인터넷 바로가기(.url) 링크 & 가격 추출기")
        self.root.geometry("550x330")  # 라디오 버튼 배치를 위해 창 높이 최적화
        self.root.resizable(False, False)

        self.selected_path = tk.StringVar()
        self.exclude_words = tk.StringVar()
        
        # 옵션 상품 처리 방식을 선택하기 위한 변수 (기본값: 자동 모드)
        self.option_mode = tk.StringVar(value="auto")

        self.create_widgets()

    def create_widgets(self):
        # 1. 폴더 선택 섹션
        folder_frame = tk.LabelFrame(self.root, text=" 1. 대상 폴더 지정 ", padx=10, pady=10)
        folder_frame.pack(fill="x", padx=15, pady=5)

        self.entry_path = tk.Entry(
            folder_frame, textvariable=self.selected_path, width=45, state="readonly"
        )
        self.entry_path.pack(side="left", padx=(0, 5), expand=True, fill="x")

        btn_browse = tk.Button(
            folder_frame, text="폴더 찾기", command=self.browse_folder, width=10
        )
        btn_browse.pack(side="right")

        # 2. 필터링 섹션
        filter_frame = tk.LabelFrame(
            self.root, text=" 2. 제외할 하위 폴더 키워드 ", padx=10, pady=10
        )
        filter_frame.pack(fill="x", padx=15, pady=5)

        entry_filter = tk.Entry(
            filter_frame, textvariable=self.exclude_words, width=58
        )
        entry_filter.pack(fill="x", ipady=2)
        entry_filter.insert(0, "품절, 제외, 보류")

        # 3. 옵션 상품 수집 모드 선택 섹션
        mode_frame = tk.LabelFrame(
            self.root, text=" 3. 옵션 상품 원가 수집 방식 설정 ", padx=10, pady=10
        )
        mode_frame.pack(fill="x", padx=15, pady=5)

        rad_auto = tk.Radiobutton(
            mode_frame, 
            text="자동 모드 (옵션 무시, 화면에 보이는 기본 최저가 즉시 수집)", 
            variable=self.option_mode, 
            value="auto"
        )
        rad_auto.pack(anchor="w")

        rad_manual = tk.Radiobutton(
            mode_frame, 
            text="반자동 모드 (선택 감지 즉시 다음 이동, 미선택 시 최대 60초 대기)", 
            variable=self.option_mode, 
            value="manual"
        )
        rad_manual.pack(anchor="w")

        # 4. 실행 버튼 섹션
        btn_start = tk.Button(
            self.root,
            text="폐쇄몰 로그인 및 링크·단가 일괄 추출",
            font=("맑은 고딕", 11, "bold"),
            bg="#2ecc71",
            fg="white",
            command=self.start_extraction,
            height=2
        )
        btn_start.pack(fill="x", padx=15, pady=10)

    def browse_folder(self):
        folder = filedialog.askdirectory(title="추출할 최상위 폴더를 선택하세요")
        if folder:
            self.selected_path.set(os.path.abspath(folder))

    def extract_url_from_file(self, file_path):
        encodings = ["cp949", "utf-8", "utf-16", "utf-8-sig", "latin-1"]
        for encoding in encodings:
            try:
                with open(file_path, "r", encoding=encoding, errors="ignore") as f:
                    for line in f:
                        clean_line = line.strip()
                        if clean_line.upper().startswith("URL="):
                            parts = clean_line.split("=", 1)
                            if len(parts) > 1:
                                return parts[1].strip()
            except Exception:
                continue
        return None

    def start_extraction(self):
        base_dir = self.selected_path.get()
        current_mode = self.option_mode.get()  # 사용자가 라디오 버튼으로 고른 모드 값

        if not base_dir:
            messagebox.showwarning("경고", "먼저 링크를 추출할 폴더를 지정해 주세요.")
            return

        raw_input = self.exclude_words.get()
        exclude_list = [w.strip() for w in raw_input.split(",") if w.strip()]

        # ---- [웹 브라우저 제어 및 반자동 로그인 대기] ----
        messagebox.showinfo(
            "로그인 안내",
            "확인을 누르면 크롬 브라우저가 열립니다.\n\n"
            "사이트 가입 아이디로 로그인을 완전히 완료하신 후에,\n"
            "다시 이 프로그램 창으로 돌아와 [확인]을 눌러주세요.",
        )

        try:
            options = webdriver.ChromeOptions()
            options.add_experimental_option("excludeSwitches", ["enable-automation"])
            options.add_experimental_option("useAutomationExtension", False)
            
            driver = webdriver.Chrome(options=options)
            driver.get("https://jy45321.imweb.me") 
        except Exception as e:
            messagebox.showerror("브라우저 실행 오류", f"크롬 브라우저를 실행할 수 없습니다.\n{e}")
            return

        messagebox.showinfo("대기 중", "로그인을 마쳤다면 이 창의 [확인]을 누르세요. 추출을 시작합니다.")

        wb = Workbook()
        
        # 1) 메인 시트 생성 및 초기화
        ws = wb.active
        ws.title = "링크 목록"
        
        # 2) 설정 시트 추가 생성 및 변수 고정 기본값 배치
        ws_config = wb.create_sheet(title="설정")
        ws_config.append(["환경변수 항목", "설정값"])
        ws_config.append(["기본 포장비", 0])          # 설정!$B$2 위치
        ws_config.append(["기본 수수료(%)", 0.00])    # 설정!$B$3 위치
        ws_config.append(["기본 판매처", "당근마켓"]) # 설정!$B$4 위치
        
        # ★ [추가] 설정 시트에 부가세 산출 공식을 시각적/관리용 항목으로 추가 (설정!$B$5 위치)
        ws_config.append(["기본 부가세 수식", "=(판매가+배송비)*10%-(원가+나의 배송비+포장비+수수료)*10%"])
        
        # 셀 서식 지정 (수수료 퍼센트 포맷팅)
        ws_config.cell(row=3, column=2).number_format = '0.00%' 
        
        # 열 매핑 가이드 (A~O)
        headers = [
            "지역", "카테고리", "상품명", "판매가", "배송비", "공급처", 
            "원가", "나의 배송비", "포장비", "판매처", "수수료(%)", 
            "수수료", "부가세", "마진", "마진율"
        ]
        ws.append(headers)

        count = 0
        url_file_found = 0

        # 데이터 탐색 및 수집
        for root_dir, dirs, files in os.walk(base_dir):
            if any(word in root_dir for word in exclude_list):
                continue

            for file in files:
                if file.lower().endswith(".url"):
                    url_file_found += 1
                    file_path = os.path.join(root_dir, file)
                    url = self.extract_url_from_file(file_path)

                    if url and "jy45321.imweb.me" in url.lower():
                        rel_path = os.path.relpath(root_dir, base_dir)
                        category_name = "최상위 폴더" if rel_path == "." else rel_path.replace(os.sep, " > ")
                        
                        product_name = os.path.splitext(file)[0]
                        hyperlink_formula = f'=HYPERLINK("{url}", "{url}")'

                        cost_val = ""         
                        delivery_fee_val = "" 
                        
                        try:
                            driver.get(url)
                            
                            # ---- [반자동 모드: 인공지능형 가변 대기 로직] ----
                            if current_mode == "manual":
                                time.sleep(1.0)  
                                
                                has_options = False
                                try:
                                    view_container = driver.find_elements(By.CSS_SELECTOR, "#shop_view, .shop_view, .shop-view")
                                    if view_container:
                                        if view_container[0].find_elements(By.CSS_SELECTOR, "select, .option_select, .opt_block, [class*='option']"):
                                            has_options = True
                                    else:
                                        has_options = True
                                except:
                                    has_options = True

                                max_wait = 60.0 if has_options else 0.8
                                start_time = time.time()
                                
                                while time.time() - start_time < max_wait:
                                    page_text = driver.find_element(By.TAG_NAME, "body").text
                                    temp_cost = ""
                                    
                                    if "총 상품금액" in page_text:
                                        cost_part = page_text.split("총 상품금액", 1)[1]
                                        cost_match = re.search(r'\(\d+개\)\s*([\d,]+)', cost_part)
                                        if cost_match:
                                            temp_cost = cost_match.group(1).replace(",", "").strip()
                                        else:
                                            cost_match = re.search(r'[\d,]+', cost_part)
                                            if cost_match:
                                                temp_cost = cost_match.group().replace(",", "").strip()
                                    
                                    if temp_cost and temp_cost != "0" and temp_cost.isdigit():
                                        break
                                    
                                    time.sleep(0.5)
                            else:
                                time.sleep(1.8) 
                            # ---- [가변 대기 로직 끝] ----
                            
                            page_text = driver.find_element(By.TAG_NAME, "body").text
                            
                            if "총 상품금액" in page_text:
                                cost_part = page_text.split("총 상품금액", 1)[1]
                                cost_match = re.search(r'\(\d+개\)\s*([\d,]+)', cost_part)
                                if cost_match:
                                    cost_val = cost_match.group(1).replace(",", "").strip()
                                else:
                                    re_match = re.search(r'[\d,]+', cost_part)
                                    if re_match:
                                        cost_val = re_match.group().replace(",", "").strip()
                            
                            if not cost_val or cost_val == "0":
                                try:
                                    price_element = driver.find_element(By.CSS_SELECTOR, "span.shop_item_price")
                                    cost_val = price_element.text.replace("원", "").replace(",", "").strip()
                                except:
                                    pass

                            if "기본" in page_text:
                                delivery_part = page_text.split("기본", 1)[1].strip()
                                target_area = delivery_part[:15]
                                
                                if "무료" in target_area:
                                    delivery_fee_val = 0  
                                else:
                                    deliv_match = re.search(r'[\d,]+', target_area)
                                    if deliv_match:
                                        delivery_fee_val = deliv_match.group().replace(",", "").strip()
                                        
                        except Exception:
                            cost_val = "오류(재확인)"
                            delivery_fee_val = "오류"

                        def convert_to_numeric(val):
                            if val is None or val == "" or str(val).lower() == "오류":
                                return 0
                            val_str = str(val).strip()
                            if val_str.isdigit():
                                return int(val_str)
                            return val

                        cost_num = convert_to_numeric(cost_val)
                        delivery_num = convert_to_numeric(delivery_fee_val)

                        # 행 인덱스 계산 (데이터는 2번째 줄부터 시작)
                        r = count + 2

                        # 설정 시트 절대 참조 수식 연결 연동
                        config_packing = "=설정!$B$2"      
                        config_market = "=설정!$B$4"       
                        config_fee_rate = "=설정!$B$3"     

                        fee_formula = f"=D{r}*K{r}+E{r}*0"
                        
                        # ★ 요청하신 수식이 메인 시트의 각 행에 맞게 동적 적용됩니다.
                        # D(판매가) + E(배송비) - G(원가) - H(나의 배송비) - I(포장비) - L(수수료)
                        vat_formula = f"=(D{r}+E{r})*10%-(G{r}+H{r}+I{r}+L{r})*10%"
                        
                        margin_formula = f"=D{r}+E{r}-G{r}-H{r}-I{r}-M{r}-L{r}"
                        margin_rate_formula = f"=IF(D{r}>0, N{r}/D{r}, 0)"

                        row_data = [
                            "",                  # 지역 (공란)
                            category_name,       # 카테고리
                            product_name,        # 상품명
                            "",                  # 판매가 (공란)
                            delivery_num,        # 배송비
                            hyperlink_formula,   # 공급처
                            cost_num,            # 원가
                            delivery_num,        # 나의 배송비 (=배송비와 동일)
                            config_packing,      # 포장비 수식 연동
                            config_market,       # 판매처 수식 연동
                            config_fee_rate,     # 수수료(%) 수식 연동
                            fee_formula,         # 수수료 결과 수식
                            vat_formula,         # 부가세 결과 수식 (요청하신 공식 빌드 완료)
                            margin_formula,      # 마진 결과 수식
                            margin_rate_formula  # 마진율 결과 수식
                        ]
                        ws.append(row_data)

                        # 퍼센트(%) 형식 표시 설정
                        ws.cell(row=r, column=11).number_format = '0.00%' 
                        ws.cell(row=r, column=15).number_format = '0.00%' 
                        
                        count += 1

        driver.quit()  

        if count > 0:
            try:
                header_fill = PatternFill(start_color="EFEFEF", end_color="EFEFEF", fill_type="solid")
                header_font = Font(name="맑은 고딕", size=11, bold=True)
                header_alignment = Alignment(horizontal="center", vertical="center")

                for col_idx in range(1, len(headers) + 1):
                    cell = ws.cell(row=1, column=col_idx)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment

                for col_idx in range(1, len(headers) + 1):
                    max_len = 0
                    col_letter = get_column_letter(col_idx)
                    for row_idx in range(1, ws.max_row + 1):
                        cell_value = ws.cell(row=row_idx, column=col_idx).value
                        if cell_value:
                            val_str = str(cell_value)
                            if val_str.startswith("="):
                                val_str = "https://example.com"
                            byte_len = len(val_str.encode("utf-8"))
                            calc_len = (byte_len - len(val_str)) / 2 + len(val_str)
                            if calc_len > max_len:
                                max_len = calc_len
                    ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

                # 설정 시트 가독성을 위해 열 너비 최적화 (수식이 길어 B열을 넓게 확보)
                ws_config.column_dimensions['A'].width = 20
                ws_config.column_dimensions['B'].width = 65

                output_path = os.path.join(base_dir, "Link_List.xlsx")
                file_counter = 1
                while True:
                    try:
                        wb.save(output_path)
                        break
                    except PermissionError:
                        output_path = os.path.join(base_dir, f"Link_List ({file_counter}).xlsx")
                        file_counter += 1

                messagebox.showinfo(
                    "성공", 
                    f"총 {count}개의 데이터 추출 완료!\n\n저장 경로: {os.path.basename(output_path)}\n\n확인을 누르면 계산서 서식 파일이 자동 열림 처리됩니다."
                )
                os.startfile(output_path)
            except Exception as e:
                messagebox.showerror("엑셀 저장 오류", f"엑셀 파일을 최종 마무리하는 도중 예기치 못한 에러가 발생했습니다:\n{e}")
        else:
            messagebox.showwarning(
                "실패", 
                f"스캔된 총 {url_file_found}개의 .url 파일 중 타겟 폐쇄몰 도메인 주소(jy45321.imweb.me)와 일치하는 유효 링크가 없거나 주소를 읽지 못했습니다."
            )


# 프로그램 정식 시작 지점 선언문
if __name__ == "__main__":
    root = tk.Tk()
    app = LinkExtractorApp(root)
    root.mainloop()
